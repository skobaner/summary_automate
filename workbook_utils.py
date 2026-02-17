from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

DATA_DIR = "data"


def find_summary_workbook(data_dir: str | Path = DATA_DIR) -> Path:
    """Return the first Excel workbook in data_dir containing 'summary' in the filename."""
    data_path = Path(data_dir)
    if not data_path.exists():
        raise FileNotFoundError(f"Data directory not found: {data_path.resolve()}")

    candidates = sorted(
        path for path in data_path.glob("*.xls*") if "summary" in path.name.lower()
    )
    if not candidates:
        raise FileNotFoundError(f"No Summary Excel files found in {data_path}")

    return candidates[0]


def extract_itb_number(filename: str) -> str:
    """Extract ITB number from strings like 'ITB15 - Summary R0.xlsx' and return unpadded text."""
    match = re.search(r"\bITB\s*0*(\d+)\b", filename, flags=re.IGNORECASE)
    if not match:
        raise ValueError(f"Could not extract ITB number from filename: {filename}")
    return match.group(1)


def extract_itb_number_from_workbook(workbook_path: str | Path) -> str:
    """
    Infer ITB number from sheet names like:
      - Summary-ITB15 (AFP BILLING)
      - ITB4-15
    """
    workbook = load_workbook(filename=Path(workbook_path), read_only=True, data_only=True)
    patterns = [
        re.compile(r"Summary-ITB\s*0*(\d+)\b", flags=re.IGNORECASE),
        re.compile(r"\bITB4-\s*0*(\d+)\b", flags=re.IGNORECASE),
    ]
    try:
        for sheet_name in workbook.sheetnames:
            for pattern in patterns:
                match = pattern.search(sheet_name)
                if match:
                    return match.group(1)
    finally:
        workbook.close()
    raise ValueError(f"Could not infer ITB number from workbook sheets: {workbook_path}")
